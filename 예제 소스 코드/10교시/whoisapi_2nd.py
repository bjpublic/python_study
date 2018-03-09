import json
import requests
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
 
# 워크북을 하나 만듭니다.
wb = Workbook()
 
# 활성화된 엑셀 시트를 선택합니다.
ws = wb.active

# 엑셀 제목을 지정합니다.
ws.title = "whois ip info"

# 각 IP 의 조회 결과를 엑셀에 넣을 때 사용한 전역 변수 입니다.
row_num = 2
 

# 제일 위쪽의 제목 줄을 저장하는 함수 입니다.
def save_field(field_list):
    # 컬럼 초기화 값
    column_char = 'a'
    
    # 리스트의 내용들을 가져와서 엑셀 맨 첫 줄에 저장합니다.
    for field in field_list:
        ws[column_char + '1'] = field
        column_char = chr(ord(column_char) + 1)
        
 
# 엑셀에 조회결과를 저장하는 함수 입니다.
def save_content(address_list):  
    # 컬럼 초기화 값
    column_char = 'a' 
    
    # 엑셀 두 번째 줄부터 IP정보를 이어서 저장하기 위해서 전역 변수를 가져옵니다.
    global row_num 
    
    # 리스트의 내용을 가져와서 row_num 을 기준으로 저장 합니다.
    for address in address_list:
        ws[column_char + str(row_num)] = address
        column_char = chr(ord(column_char) + 1)





# IP 하나에 대해 WHOIS API 를 호출하는 함수 입니다.
def requestWhois(searchIP):
    # 엑셀 2번째 줄부터 IP정보를 이어서 저장하기 위해서 전역 변수를 가져옵니다.
    global row_num 

    # 결과를 담을 배열을 초기화 합니다.
    whois_data = []
  
    # API 를 호출 합니다(발급받은키는 독자분의 키로 수정해야 합니다)
    con = s.get('http://whois.kisa.or.kr/openapi/whois.jsp?query=' + searchIP + '&key=발급받은키&answer=json')
 
    # 호출된 결과를 json 형태로 파싱하여 가져옵니다.
    json_data = json.loads(con.text)

    # 각 결과 값을 리스트에 추가 합니다.
    whois_data.append(json_data['whois']['query'])
    whois_data.append(json_data['whois']['countryCode'])
 
    # 해당 값들이 없어 에러처리가 나면 무시 합니다.
    try:
        whois_data.append(json_data['whois']['korean']['PI']['netinfo']['addr'])
        whois_data.append(json_data['whois']['korean']['PI']['netinfo']['range'])
        whois_data.append(json_data['whois']['korean']['PI']['netinfo']['servName'])
    except:
        pass
    
    # 해당 값들이 없어 에러처리가 나면 무시 합니다.    
    try:
        whois_data.append(json_data['whois']['korean']['ISP']['netinfo']['addr'])
        whois_data.append(json_data['whois']['korean']['ISP']['netinfo']['range'])
        whois_data.append(json_data['whois']['korean']['ISP']['netinfo']['servName'])
    except:
        pass

    # 결과를 로그 용으로 화면에 뿌립니다.
    print(whois_data)
 
    # 엑셀에 결과를 저장합니다.
    save_content(whois_data)
    row_num = row_num + 1
  
 
# 엑셀 첫 째 줄에 각 필드들의  제목을 저장합니다.
excel_fields = ['IP', 'countryCode', 'addr', 'IP_range', 'servName']
save_field(excel_fields)
 
# IP 가 담긴 텍스트 파일을 열어 줄로 나눠서 리스트에 담습니다.
f = open('ip.txt', 'r')
iplist = f.read().splitlines()

# 텍스트 파일을 닫습니다.
f.close()

# API 를 호출할 세션을 생성합니다.
s = requests.session()

# IP 별로 조회하여 엑셀에 저장합니다.
for ip in iplist:
  requestWhois(ip)

# 파일을 실제 저장 합니다.
wb.save("ipinfo.xlsx")