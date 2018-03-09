from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
import pyodbc

# 워크북을 하나 만듭니다.
wb = Workbook()
 
# 활성화된 엑셀 시트를 선택합니다.
ws = wb.active

# 엑셀 제목을 지정합니다.
ws.title = "output"
 
# 데이터베이스 연결 문자열을 세팅 합니다.
server = 'localhost'
database = 'mytest'
username = 'pyuser'
password = 'test1234'

# 데이터 베이스에 연결 합니다.
cnxn = pyodbc.connect('DRIVER={ODBC Driver 13 for SQL Server};SERVER='+server+';PORT=1433;DATABASE='+database+';UID='+username+';PWD='+ password)

# 커서를 만듭니다.
cursor = cnxn.cursor()

# 가져올 테이블 이름 변수 입니다.
my_table_name = 'supermarket'


# A) 칼럼이름 얻어오기 함수 입니다.
def get_column_name(table_name):
    column_name = []
    
    cursor.execute('SELECT column_name FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = \'' + table_name + '\';')

    row = cursor.fetchone()  
    while row:
        column_name.append(row[0])
        row = cursor.fetchone()
    return column_name

 
# B) 칼럼 이름을 엑셀에 저장하는 함수 입니다.
def save_column_name(column_name):
    column_char = 'a'

    for name in column_name:
        ws[column_char + '1'] = name
        column_char = chr(ord(column_char) + 1)

 
# C) select 칼럼 쿼리 제작하는 함수 입니다(Item, Category, ..., Price)
def make_column_query(column_name):
    column_query = ''

    for name in column_name:
        column_query = column_query + name + ','
    column_query = column_query[:-1]
    return column_query

 
# D) 테이블 내용을 엑셀에 저장하는 함수 입니다.
def save_table_content(table_name, column_name):
    column_query = make_column_query(column_name)

    cursor.execute('SELECT ' + column_query + ' FROM ' + table_name + '(nolock);')
    row = cursor.fetchone()
 
    row_num = 2
    while row:
        column_char = 'a'  
        for x in range(1, len(column_name)+1):  #칼럼수 참조하게 변경
            ws[column_char + str(row_num)] = row[x-1]
            column_char = chr(ord(column_char) + 1) 

        row_num = row_num + 1
        row = cursor.fetchone()
 
### 만든 함수들 실행
# 칼럼 이름을  얻어옵니다. 
my_column_name = get_column_name(my_table_name)
# 칼럼 이름을 저장합니다.
save_column_name(my_column_name)
# 테이블 내용을 저장합니다.
save_table_content(my_table_name, my_column_name)
  
# 파일을 실제 저장 합니다.
wb.save("result2.xlsx")

print("엑셀저장 완료")