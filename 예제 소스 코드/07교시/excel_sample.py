from openpyxl import Workbook
import datetime
 
# 워크북을 하나 만듭니다.
wb = Workbook()
 
# 활성 워크시트를 선택 합니다(첫 번째 시트가 선택 됩니다).
ws = wb.active
 
# A1 행에 42라는 숫자를 넣습니다.
ws['A1'] = 42
 
# 현재 글자가 쓰여있는 다음 row 에 1,2,3 이라고 넣습니다.
ws.append([1, 2, 3])
 
# 셀이 겹치는 부분을 수정 했습니다.
ws['A3'] = datetime.datetime.now()
 
# 메모리에 있는 워크 북을 실제 물리적인 엑셀 파일로 저장한다.
wb.save("sample.xlsx")

print("엑셀저장 완료")