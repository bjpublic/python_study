from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
import pyodbc
 
# 워크북을 하나 만듭니다.
wb = Workbook()
 
# 활성화된 엑셀 시트를 선택합니다.
ws = wb.active

# 엑셀 제목을 지정합니다.
ws.title = "output"


# 데이터베이스 연결 문자열을 세팅 합니다.
server = 'localhost'
database = 'mytest'
username = 'pyuser'
password = 'test1234'

# 데이터 베이스에 연결 합니다.
cnxn = pyodbc.connect('DRIVER={ODBC Driver 13 for SQL Server};SERVER='+server+';PORT=1433;DATABASE='+database+';UID='+username+';PWD='+ password)

# 커서를 만듭니다.
cursor = cnxn.cursor()


### 칼럼 이름 저장하기 
# supermarket 의 칼럼들을 가져옵니다.
cursor.execute('SELECT column_name FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = \'supermarket\';')

# 첫 번째 칼럼을 나타냅니다. 
column_char = 'a'
 
# 한 행씩 가져오면서,
row = cursor.fetchone()
while row:
    # 칼럼 문자를 하나씩 증가하면서 A1~E1에 조회해 온 칼럼 명을 넣습니다.
    ws[column_char + '1'] = row[0]
    column_char = chr(ord(column_char) + 1)
    row = cursor.fetchone()


### 테이블 내용 저장 하기
# supermarket 테이블의 내용을 가져옵니다.
cursor.execute('SELECT Itemno, Category, FoodName, Company, Price FROM supermarket(nolock);')

# 2번째 행을 나타냅니다.
row_num = 2
 
# 한 행씩 가져오면서,
row = cursor.fetchone()
while row:
    # 예전 수동 타자기 처럼, 새로운 줄이 오게 되면, 첫째 열인 A 로 돌아가는 초기 값 입니다.
    column_char = 'a' 

    # 1~5 까지 x 가 변하면서 칼럼 문자, row를 하나씩 늘여 결과를 하나씩 담습니다. 
    # ws['a1'] = row[0], ws['b1'] = row[1], ws['c1'] = row[2]...
    for x in range(1, 6):
        ws[column_char + str(row_num)] = row[x-1]
        column_char = chr(ord(column_char) + 1)
 
    # 다음 행을 표시하기 위해 뒤의 숫자를 증가 시킵니다.       
    row_num = row_num + 1
    row = cursor.fetchone()

 
# 파일을 실제 저장 합니다.
wb.save("result.xlsx")

print("엑셀저장 완료")
